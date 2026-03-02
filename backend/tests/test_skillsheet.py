"""
Phase 7 スキルシート出力 API テスト

- POST /api/v1/skillsheet/export
- GET  /api/v1/skillsheet/download/{token}
"""
import io
import uuid
import zipfile
import pytest
from datetime import date, datetime, UTC

from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import create_access_token
from app.models.employee import Employee


# ── ローカルフィクスチャ ─────────────────────────────────────────────────────


@pytest.fixture
async def sales_employee(db: AsyncSession) -> Employee:
    emp = Employee(
        id=str(uuid.uuid4()),
        employee_number="TEST-SALES",
        name_ja="テスト営業",
        name_en="Test Sales",
        system_role="sales",
        office_location="HANOI",
        employment_type="FULLTIME",
        work_style="REMOTE",
        joined_at=date(2024, 1, 1),
        email="sales@test.local",
        is_active=True,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )
    db.add(emp)
    await db.flush()
    return emp


@pytest.fixture
def sales_headers(sales_employee: Employee) -> dict:
    token = create_access_token(
        subject=sales_employee.id,
        extra={"role": sales_employee.system_role},
    )
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def second_employee(db: AsyncSession) -> Employee:
    """2人目の社員（combined/zip テスト用）。"""
    emp = Employee(
        id=str(uuid.uuid4()),
        employee_number="TEST-003",
        name_ja="テスト社員2",
        name_en="Test Employee 2",
        system_role="member",
        office_location="HCMC",
        employment_type="FULLTIME",
        work_style="ONSITE",
        joined_at=date(2024, 3, 1),
        email="emp2@test.local",
        is_active=True,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )
    db.add(emp)
    await db.flush()
    return emp


# ── 1. export 200 admin ────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_admin_ok(client: AsyncClient, admin_employee: Employee, auth_headers: dict):
    """管理者は export に成功する（200）。"""
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [admin_employee.id]},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert "download_url" in data
    assert "expires_at" in data
    assert "filename" in data


# ── 2. export 200 sales ────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_sales_ok(client: AsyncClient, sales_employee: Employee, sales_headers: dict, admin_employee: Employee):
    """sales ロールも export に成功する（200）。"""
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [admin_employee.id]},
        headers=sales_headers,
    )
    assert resp.status_code == 200


# ── 3. export 403 member ───────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_member_forbidden(client: AsyncClient, member_employee: Employee):
    """member ロールは export が拒否される（403）。"""
    token = create_access_token(
        subject=member_employee.id,
        extra={"role": member_employee.system_role},
    )
    headers = {"Authorization": f"Bearer {token}"}
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [member_employee.id]},
        headers=headers,
    )
    assert resp.status_code == 403


# ── 4. export 空リスト → 422 ───────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_empty_ids_422(client: AsyncClient, admin_employee: Employee, auth_headers: dict):
    """employee_ids が空リストの場合は 422 バリデーションエラー。"""
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": []},
        headers=auth_headers,
    )
    assert resp.status_code == 422


# ── 5. export 存在しない employee_id → 404 ─────────────────────────────────

@pytest.mark.asyncio
async def test_export_nonexistent_employee_404(client: AsyncClient, auth_headers: dict):
    """存在しない employee_id を指定した場合は 404。"""
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [str(uuid.uuid4())]},
        headers=auth_headers,
    )
    assert resp.status_code == 404


# ── 6. export xlsx → download_url に token 含む ────────────────────────────

@pytest.mark.asyncio
async def test_export_xlsx_download_url(client: AsyncClient, admin_employee: Employee, auth_headers: dict):
    """xlsx エクスポートで download_url が /api/v1/skillsheet/download/<token> 形式。"""
    resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [admin_employee.id], "format": "xlsx"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    url: str = data["download_url"]
    assert url.startswith("/api/v1/skillsheet/download/")
    # URL の最後のセグメントが UUID 形式であることを確認
    token_part = url.split("/")[-1]
    assert len(token_part) == 36  # UUID 長


# ── 7. download 有効 token → 200 + xlsx Content-Type ──────────────────────

@pytest.mark.asyncio
async def test_download_valid_token(client: AsyncClient, admin_employee: Employee, auth_headers: dict):
    """有効なトークンでダウンロードすると 200 + xlsx の Content-Type を返す。"""
    # まずエクスポートしてトークンを取得
    export_resp = await client.post(
        "/api/v1/skillsheet/export",
        json={"employee_ids": [admin_employee.id], "format": "xlsx"},
        headers=auth_headers,
    )
    assert export_resp.status_code == 200
    token = export_resp.json()["download_url"].split("/")[-1]

    # ダウンロード（認証不要）
    dl_resp = await client.get(f"/api/v1/skillsheet/download/{token}")
    assert dl_resp.status_code == 200
    assert "spreadsheetml" in dl_resp.headers["content-type"]


# ── 8. download 存在しない token → 404 ────────────────────────────────────

@pytest.mark.asyncio
async def test_download_invalid_token_404(client: AsyncClient):
    """存在しないトークンで 404 が返る。"""
    resp = await client.get(f"/api/v1/skillsheet/download/{uuid.uuid4()}")
    assert resp.status_code == 404


# ── 9. combined + 2名 → 1ファイル（2シート）──────────────────────────────

@pytest.mark.asyncio
async def test_export_combined_two_employees(
    client: AsyncClient,
    admin_employee: Employee,
    second_employee: Employee,
    auth_headers: dict,
):
    """combined スタイルで2名 → 2シート入りの xlsx が1ファイル生成される。"""
    export_resp = await client.post(
        "/api/v1/skillsheet/export",
        json={
            "employee_ids": [admin_employee.id, second_employee.id],
            "format": "xlsx",
            "output_style": "combined",
        },
        headers=auth_headers,
    )
    assert export_resp.status_code == 200
    token = export_resp.json()["download_url"].split("/")[-1]

    dl_resp = await client.get(f"/api/v1/skillsheet/download/{token}")
    assert dl_resp.status_code == 200

    # openpyxl で読み込んでシート数を確認
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(dl_resp.content))
    assert len(wb.sheetnames) == 2


# ── 10. zip + 2名 → zip ファイル ──────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_zip_two_employees(
    client: AsyncClient,
    admin_employee: Employee,
    second_employee: Employee,
    auth_headers: dict,
):
    """zip スタイルで2名 → ZIP ファイルが返り、内部に2ファイルある。"""
    export_resp = await client.post(
        "/api/v1/skillsheet/export",
        json={
            "employee_ids": [admin_employee.id, second_employee.id],
            "format": "xlsx",
            "output_style": "zip",
        },
        headers=auth_headers,
    )
    assert export_resp.status_code == 200

    token = export_resp.json()["download_url"].split("/")[-1]
    dl_resp = await client.get(f"/api/v1/skillsheet/download/{token}")
    assert dl_resp.status_code == 200
    assert dl_resp.headers["content-type"] == "application/zip"

    # ZIP 内のファイル数を確認
    with zipfile.ZipFile(io.BytesIO(dl_resp.content)) as zf:
        assert len(zf.namelist()) == 2
