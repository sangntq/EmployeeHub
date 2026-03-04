"""
NJP形式スキルシート生成サービス

- Excel (.xlsx): openpyxl を使用して職務経歴書フォーマット（NJP仕様）
- PDF          : reportlab を使用して同等フォーマット
- combined     : 複数社員を1ファイルにまとめる
- zip          : 社員ごとの個別ファイルを ZIP にまとめる
"""
import io
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.employee import Employee
from app.models.skill import EmployeeSkill
from app.models.certification import EmployeeCertification
from app.models.project import EmployeeProject

EXPORTS_DIR = Path("exports")

# 担当工程の標準順序（NJP仕様）
PROCESS_PHASES_ORDER = [
    "要件定義", "基本設計", "詳細設計", "製造・実装",
    "単体テスト", "結合テスト", "総合テスト", "運用・保守",
]


# ── Excel 生成 ────────────────────────────────────────────────────────────────

def _write_njp_excel_sheet(ws, emp: Employee) -> None:
    """NJP形式で社員1名分の職務経歴書をワークシートに書き込む。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BLUE = "1F6DAF"
    HEADER_BG = "D9E1F2"
    SECTION_BG = "EBF0FA"
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def label_cell(row, col, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
        return cell

    def section_header(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=11, color=BLUE)
        ws.merge_cells(f"A{row}:E{row}")

    # タイトル
    ws["A1"] = "職務経歴書"
    ws["A1"].font = Font(bold=True, size=16, color=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    # 基本情報
    row = 3
    dept_name = emp.department.name_ja if emp.department else "—"
    basic_rows = [
        ("氏名",    emp.name_ja or "—",   "社員番号", emp.employee_number),
        ("部署",    dept_name,             "入社日",   str(emp.joined_at)),
        ("拠点",    emp.office_location,   "日本語Lv", emp.japanese_level or "—"),
        ("メール",  emp.email,             "",         ""),
    ]
    for label1, val1, label2, val2 in basic_rows:
        label_cell(row, 1, label1)
        ws.cell(row=row, column=2, value=val1)
        if label2:
            label_cell(row, 4, label2)
            ws.cell(row=row, column=5, value=val2)
        row += 1
    row += 1

    # 自己PR
    self_pr = getattr(emp, "self_pr", None)
    if self_pr:
        section_header(row, "【自己PR】")
        row += 1
        c = ws.cell(row=row, column=1, value=self_pr)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = max(30, min(120, len(self_pr) // 4))
        row += 2

    # 技術スキルサマリ
    section_header(row, "【技術スキルサマリ】")
    row += 1
    approved_skills = [s for s in emp.skills if s.status == "APPROVED"]
    if approved_skills:
        for col, h in enumerate(["スキル名", "承認Lv", "経験年数（年）", "最終使用"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
            cell.border = thin_border
        row += 1
        for s in approved_skills:
            ws.cell(row=row, column=1, value=s.skill.name if s.skill else "—").border = thin_border
            ws.cell(row=row, column=2, value=s.approved_level).border = thin_border
            ws.cell(row=row, column=3, value=float(s.experience_years) if s.experience_years else None).border = thin_border
            ws.cell(row=row, column=4, value=str(s.last_used_at) if s.last_used_at else "—").border = thin_border
            row += 1
    else:
        ws.cell(row=row, column=1, value="（スキルなし）")
        row += 1
    row += 1

    # 保有資格
    section_header(row, "【保有資格】")
    row += 1
    approved_certs = [c for c in emp.certifications if c.status == "APPROVED"]
    if approved_certs:
        for col, h in enumerate(["資格名", "カテゴリ", "取得日", "有効期限"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
            cell.border = thin_border
        row += 1
        for c in approved_certs:
            name = c.custom_name or (c.certification_master.name if c.certification_master else "—")
            cat = c.certification_master.category if c.certification_master else "—"
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value=cat).border = thin_border
            ws.cell(row=row, column=3, value=str(c.obtained_at)).border = thin_border
            ws.cell(row=row, column=4, value=str(c.expires_at) if c.expires_at else "—").border = thin_border
            row += 1
    else:
        ws.cell(row=row, column=1, value="（資格なし）")
        row += 1
    row += 1

    # 職務経歴
    section_header(row, "【職務経歴】")
    row += 1
    sorted_projects = sorted(emp.projects, key=lambda p: p.started_at, reverse=True)
    if not sorted_projects:
        ws.cell(row=row, column=1, value="（プロジェクト経歴なし）")
    else:
        for ep in sorted_projects:
            pname = ep.project.name if ep.project else "—"
            period = f"{ep.started_at}〜{ep.ended_at or '現在'}"
            client = (ep.project.client_name if ep.project else None) or "—"
            industry = (ep.project.industry if ep.project else None) or "—"

            # プロジェクト見出し
            c = ws.cell(row=row, column=1, value=f"■ {pname}")
            c.font = Font(bold=True, size=11)
            c.fill = PatternFill(fill_type="solid", fgColor=SECTION_BG)
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

            label_cell(row, 1, "期間")
            ws.cell(row=row, column=2, value=period)
            label_cell(row, 3, "顧客")
            ws.cell(row=row, column=4, value=client)
            row += 1

            label_cell(row, 1, "業種")
            ws.cell(row=row, column=2, value=industry)
            label_cell(row, 3, "規模")
            ws.cell(row=row, column=4, value=f"{ep.team_size}名" if ep.team_size else "—")
            row += 1

            label_cell(row, 1, "担当ロール")
            ws.cell(row=row, column=2, value=ep.role)
            row += 1

            # 担当工程（■/□チェックボックス）
            ep_phases = getattr(ep, "process_phases", None) or []
            phase_text = "  ".join(
                f"■{ph}" if ph in ep_phases else f"□{ph}"
                for ph in PROCESS_PHASES_ORDER
            )
            label_cell(row, 1, "担当工程")
            c = ws.cell(row=row, column=2, value=phase_text)
            c.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row}:E{row}")
            row += 1

            tech = ", ".join(ep.tech_stack) if ep.tech_stack else "—"
            label_cell(row, 1, "使用技術")
            c = ws.cell(row=row, column=2, value=tech)
            c.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row}:E{row}")
            row += 1

            if ep.responsibilities:
                label_cell(row, 1, "担当業務")
                c = ws.cell(row=row, column=2, value=ep.responsibilities)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"B{row}:E{row}")
                ws.row_dimensions[row].height = max(30, min(90, len(ep.responsibilities) // 3))
                row += 1

            lessons = getattr(ep, "lessons_learned", None)
            if lessons:
                label_cell(row, 1, "習得スキル")
                c = ws.cell(row=row, column=2, value=lessons)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"B{row}:E{row}")
                ws.row_dimensions[row].height = max(30, min(90, len(lessons) // 3))
                row += 1

            row += 1  # プロジェクト間の空白行

    col_widths = [14, 32, 10, 22, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_xlsx(employees: list[Employee]) -> bytes:
    """社員リストから Excel ワークブック（複数シート）を生成してバイト列を返す。"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for emp in employees:
        sheet_name = (emp.name_ja or emp.employee_number)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_njp_excel_sheet(ws, emp)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF 生成 ──────────────────────────────────────────────────────────────────

def generate_pdf(employees: list[Employee]) -> bytes:
    """社員リストから NJP 形式 PDF を生成してバイト列を返す。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Table, TableStyle,
        Spacer, PageBreak,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    FONT = "HeiseiKakuGo-W5"
    BLUE = colors.HexColor("#1F6DAF")
    HEADER_BG = colors.HexColor("#D9E1F2")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )

    title_s = ParagraphStyle("T", fontName=FONT, fontSize=16, spaceAfter=6, textColor=BLUE, alignment=1)
    head_s = ParagraphStyle("H", fontName=FONT, fontSize=10, spaceAfter=2, textColor=BLUE)
    norm_s = ParagraphStyle("N", fontName=FONT, fontSize=8, spaceAfter=2)
    proj_s = ParagraphStyle("P", fontName=FONT, fontSize=9, spaceAfter=2, textColor=BLUE)

    base_ts = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
    ])

    story: list = []

    for emp_idx, emp in enumerate(employees):
        if emp_idx > 0:
            story.append(PageBreak())

        story.append(Paragraph("職務経歴書", title_s))
        story.append(Spacer(1, 3 * mm))

        # 基本情報テーブル
        dept_name = emp.department.name_ja if emp.department else "—"
        info_data = [
            ["氏名", emp.name_ja or "—", "社員番号", emp.employee_number],
            ["部署", dept_name, "入社日", str(emp.joined_at)],
            ["拠点", emp.office_location, "日本語Lv", emp.japanese_level or "—"],
            ["メール", emp.email, "", ""],
        ]
        info_t = Table(info_data, colWidths=[22 * mm, 56 * mm, 22 * mm, 56 * mm])
        info_t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), HEADER_BG),
            ("BACKGROUND", (2, 0), (2, -1), HEADER_BG),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(info_t)
        story.append(Spacer(1, 4 * mm))

        self_pr = getattr(emp, "self_pr", None)
        if self_pr:
            story.append(Paragraph("【自己PR】", head_s))
            story.append(Paragraph(self_pr, norm_s))
            story.append(Spacer(1, 3 * mm))

        # 技術スキルサマリ
        story.append(Paragraph("【技術スキルサマリ】", head_s))
        approved_skills = [s for s in emp.skills if s.status == "APPROVED"]
        if approved_skills:
            skill_data = [["スキル名", "承認Lv", "経験年数", "最終使用"]]
            for s in approved_skills:
                skill_data.append([
                    s.skill.name if s.skill else "—",
                    str(s.approved_level) if s.approved_level else "—",
                    f"{float(s.experience_years):.1f}年" if s.experience_years else "—",
                    str(s.last_used_at) if s.last_used_at else "—",
                ])
            t = Table(skill_data, colWidths=[68 * mm, 20 * mm, 22 * mm, 22 * mm])
            t.setStyle(base_ts)
            story.append(t)
        else:
            story.append(Paragraph("（スキルなし）", norm_s))
        story.append(Spacer(1, 3 * mm))

        # 保有資格
        story.append(Paragraph("【保有資格】", head_s))
        approved_certs = [c for c in emp.certifications if c.status == "APPROVED"]
        if approved_certs:
            cert_data = [["資格名", "カテゴリ", "取得日", "有効期限"]]
            for c in approved_certs:
                name = c.custom_name or (c.certification_master.name if c.certification_master else "—")
                cat = c.certification_master.category if c.certification_master else "—"
                cert_data.append([name, cat, str(c.obtained_at), str(c.expires_at) if c.expires_at else "—"])
            t = Table(cert_data, colWidths=[68 * mm, 25 * mm, 20 * mm, 19 * mm])
            t.setStyle(base_ts)
            story.append(t)
        else:
            story.append(Paragraph("（資格なし）", norm_s))
        story.append(Spacer(1, 3 * mm))

        # 職務経歴
        story.append(Paragraph("【職務経歴】", head_s))
        sorted_projects = sorted(emp.projects, key=lambda p: p.started_at, reverse=True)
        if not sorted_projects:
            story.append(Paragraph("（プロジェクト経歴なし）", norm_s))
        else:
            for ep in sorted_projects:
                pname = ep.project.name if ep.project else "—"
                period = f"{ep.started_at}〜{ep.ended_at or '現在'}"
                client = (ep.project.client_name if ep.project else None) or "—"
                industry = (ep.project.industry if ep.project else None) or "—"
                ep_phases = getattr(ep, "process_phases", None) or []
                phase_text = "  ".join(
                    f"■{ph}" if ph in ep_phases else f"□{ph}"
                    for ph in PROCESS_PHASES_ORDER
                )
                tech = ", ".join(ep.tech_stack) if ep.tech_stack else "—"
                team = f"{ep.team_size}名" if ep.team_size else "—"

                story.append(Paragraph(f"■ {pname}", proj_s))

                detail_rows = [
                    ["期間", period, "顧客", client],
                    ["業種", industry, "規模", team],
                    ["担当ロール", ep.role, "", ""],
                    ["担当工程", phase_text, "", ""],
                    ["使用技術", tech, "", ""],
                ]
                if ep.responsibilities:
                    detail_rows.append(["担当業務", ep.responsibilities, "", ""])
                lessons = getattr(ep, "lessons_learned", None)
                if lessons:
                    detail_rows.append(["習得スキル", lessons, "", ""])

                detail_t = Table(detail_rows, colWidths=[22 * mm, 54 * mm, 17 * mm, 63 * mm])
                ts = TableStyle([
                    ("FONTNAME", (0, 0), (-1, -1), FONT),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("BACKGROUND", (0, 0), (0, -1), HEADER_BG),
                    ("BACKGROUND", (2, 0), (2, -1), HEADER_BG),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("WORDWRAP", (1, 0), (3, -1), True),
                ])
                # 行2以降のB〜D列を結合（担当ロール/工程/技術/業務/習得スキル）
                for r_idx in range(2, len(detail_rows)):
                    ts.add("SPAN", (1, r_idx), (3, r_idx))
                detail_t.setStyle(ts)
                story.append(detail_t)
                story.append(Spacer(1, 3 * mm))

    doc.build(story)
    return buf.getvalue()


# ── メインエクスポート処理 ─────────────────────────────────────────────────────

async def export_skillsheet(
    db: AsyncSession,
    employee_ids: list[str],
    format: str,
    output_style: str,
    filename_prefix: str,
    include_salary: bool,
) -> dict:
    """
    スキルシートを生成してファイルを保存し、ダウンロード情報を返す。

    Returns:
        dict: download_url, expires_at, filename
    """
    EXPORTS_DIR.mkdir(exist_ok=True)

    stmt = (
        select(Employee)
        .where(Employee.id.in_(employee_ids))
        .options(
            selectinload(Employee.skills).selectinload(EmployeeSkill.skill),
            selectinload(Employee.certifications).selectinload(EmployeeCertification.certification_master),
            selectinload(Employee.projects).selectinload(EmployeeProject.project),
            selectinload(Employee.department),
        )
    )
    result = await db.execute(stmt)
    employees = result.scalars().all()

    found_ids = {emp.id for emp in employees}
    missing = set(employee_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"Employees not found: {missing}")

    ext = "zip" if output_style == "zip" else ("pdf" if format == "pdf" else "xlsx")
    token = str(uuid.uuid4())
    filename = f"{filename_prefix}_{token[:8]}.{ext}"
    filepath = EXPORTS_DIR / f"{token}.{ext}"

    if output_style == "zip":
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for emp in employees:
                safe_name = (emp.name_ja or emp.employee_number).replace("/", "_")
                if format == "pdf":
                    content_bytes = generate_pdf([emp])
                    inner_name = f"{safe_name}.pdf"
                else:
                    content_bytes = generate_xlsx([emp])
                    inner_name = f"{safe_name}.xlsx"
                zf.writestr(inner_name, content_bytes)
        filepath.write_bytes(zip_buf.getvalue())
    elif format == "pdf":
        filepath.write_bytes(generate_pdf(list(employees)))
    else:
        filepath.write_bytes(generate_xlsx(list(employees)))

    expires_at = datetime.now(timezone.utc).replace(microsecond=0)
    expires_at = expires_at.replace(hour=(expires_at.hour + 1) % 24)

    return {
        "download_url": f"/api/v1/skillsheet/download/{token}",
        "expires_at": expires_at.isoformat(),
        "filename": filename,
    }
